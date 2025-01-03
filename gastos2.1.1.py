import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, NamedStyle, Alignment
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog
import pandas as pd
def configurar_aba(planilha, nome, num_linhas, num_colunas, soma_colunas, divisorias, estilo_moeda): #início função para automatizar a criação das páginas
    aba = planilha.create_sheet(nome)
    aba.sheet_view.showGridLines = False #retira linhas de grade
    #criar cabeçalhos
    aba.merge_cells('A1:D1')
    aba['A1'] = 'Cartão de crédito'
    aba['A1'].alignment = Alignment(horizontal='center', vertical='center')
    aba['A1'].font = Font(color="FFFFFF", bold=True, size=14)
    aba['A1'].fill = preto
    aba.merge_cells('F1:I1')
    aba['F1'] = 'Flash'
    aba['F1'].alignment = Alignment(horizontal='center', vertical='center')
    aba['F1'].font = Font(color="FFFFFF", bold=True, size=14)
    aba['F1'].fill = preto
    aba.merge_cells('K1:N1')
    aba['K1'] = 'Débito/Pix'
    aba['K1'].alignment = Alignment(horizontal='center', vertical='center')
    aba['K1'].font = Font(color="FFFFFF", bold=True, size=14)
    aba['K1'].fill = preto
    cabecalho = ["Compra", "Preço", "Dia", "Loja", " ", "Compra", "Preço", "Dia", "Loja", " ", "Compra", "Preço", "Dia", "Loja"]
    for col, valor in enumerate(cabecalho, start=1):
        celula = aba.cell(row=2, column=col, value=valor)
        celula.font = Font(bold=True)
        celula.fill = cinza
    aba.freeze_panes = "A4" #congela linhas 1, 2 e 3
    #formata células e textos
    for col in range(1, num_colunas + 1):
        aba.cell(row=3, column=col).fill = cinza
    for row in range(4, num_linhas + 1):
        for col in soma_colunas:
            aba.cell(row=row, column=col).style = estilo_moeda
    for row in range(1, num_linhas + 1):
        for col in range(1, num_colunas + 1):
            aba.cell(row=row, column=col).border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
    for col in divisorias: #cria divisórias
        aba.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 0.5
        for row in range(1, num_linhas + 1):
            aba.cell(row=row, column=col).fill = preto
    largura_colunas = {1: 30, 2: 11, 3: 4, 4: 30, 6: 30, 7: 11, 8: 4, 9: 30, 11: 30, 12: 11, 13: 4, 14: 30} #define largura das colunas
    for col, largura in largura_colunas.items():
        aba.column_dimensions[openpyxl.utils.get_column_letter(col)].width = largura
    aba.cell(row=3, column=1, value="Total:").font = Font(bold=True)
    aba.cell(row=3, column=6, value="Total:").font = Font(bold=True)
    aba.cell(row=3, column=11, value="Total:").font = Font(bold=True)
    for col, formula in zip(soma_colunas, soma_colunas):  #soma colunas de preço
        celula_soma = aba.cell(row=3, column=col, value=f'=SUM({openpyxl.utils.get_column_letter(col)}4:{openpyxl.utils.get_column_letter(col)}{num_linhas})')
        celula_soma.style = estilo_moeda
        celula_soma.font = Font(bold=True)
        celula_soma.fill = cinza
    return aba #fim função para automatizar a criação das páginas
planilha = openpyxl.Workbook() #criar planilha
sheet = planilha.active
planilha.remove(sheet) #exclui planilha padrão sheet
preto = PatternFill(start_color="FF000000", end_color="FF000000", fill_type="solid") #define preenchimento preto
cinza = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid") #define preenchimento cinza
moeda = NamedStyle(name="moeda")
moeda.number_format = 'R$ #,##0.00' #define formato para moeda
meses = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'] #criar abas
for mes in meses: #chama função de criação das páginas
    configurar_aba(
        planilha=planilha,
        nome=mes,
        num_linhas=50,
        num_colunas=14,
        soma_colunas=[2, 7, 12],
        divisorias=[5, 10],
        estilo_moeda=moeda
    )
def salvar_planilha(): #início função para salvar planilha
    root = tk.Tk()
    root.withdraw()
    titulo = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")],
        title="Salvar planilha como"
    )
    if titulo:
        planilha.save(titulo)
        print(f"Planilha salva em: {titulo}")
    else:
        print("Nenhum título foi selecionado.") #fim função para salvar planilha
if __name__ == "__main__": #chama função de salvar arquivo
    salvar_planilha() 