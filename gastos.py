'''
versão 1.0.0
'''

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, NamedStyle

#criar arquivo xlxs
planilha = openpyxl.Workbook()
#excluir página padrão
sheet = planilha.active
planilha.remove(sheet)
#definir bordas
contorno = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)
#definir colunas
compra1 = 1
preco1 = 2
data1 = 3
loja1 = 4
divisoria = 5
compra2 = 6
preco2 = 7
data2 = 8
loja2 = 9
#definir tamanho da tabela
num_linhas = 50
num_colunas = 9
#definir soma
total = 50 
soma_coluna_2 = f'=SUM(B2:B{total - 1})'
soma_coluna_7 = f'=SUM(G2:G{total - 1})'
#definir divisoria preta
preenchimento = PatternFill(start_color="FF000000", end_color="FF000000", fill_type="solid")
#definir formato moeda
moeda = NamedStyle(name="moeda")
moeda.number_format = 'R$ #,##0.00'
#criar páginas com linhas e colunas formatadas
janeiro = planilha.create_sheet('Janeiro')
cabecalho = ["Compra", "Preço", "Data", "Loja", " ", "Compra", "Preço", "Data", "Loja"]
cinza = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
for col, valor in enumerate(cabecalho, start=1):
    celula = janeiro.cell(row=1, column=col, value=valor)
    celula.font = Font(bold=True)
    celula.fill = cinza
#aplicar estilo moeda
for row in range(2, 51):
    celula_coluna_2 = janeiro.cell(row=row, column=2)
    celula_coluna_7 = janeiro.cell(row=row, column=7)
    celula_coluna_2.style = moeda
    celula_coluna_7.style = moeda
#aplicar bordas
for i in range(1, num_linhas + 1):
    for j in range(1, num_colunas + 1):
        cell = janeiro.cell(row=i, column=j)
        cell.border = contorno
#formatar colunas
janeiro.column_dimensions[openpyxl.utils.get_column_letter(divisoria)].width = 0.5
janeiro.column_dimensions[openpyxl.utils.get_column_letter(compra1)].width = 30
janeiro.column_dimensions[openpyxl.utils.get_column_letter(loja1)].width = 30
janeiro.column_dimensions[openpyxl.utils.get_column_letter(compra2)].width = 30
janeiro.column_dimensions[openpyxl.utils.get_column_letter(loja2)].width = 30
janeiro.column_dimensions[openpyxl.utils.get_column_letter(preco1)].width = 11
janeiro.column_dimensions[openpyxl.utils.get_column_letter(preco2)].width = 11
janeiro.column_dimensions[openpyxl.utils.get_column_letter(data1)].width = 5
janeiro.column_dimensions[openpyxl.utils.get_column_letter(data2)].width = 5
for row in range(1, 51):
    celula = janeiro.cell(row=row, column=divisoria)
    celula.fill = preenchimento
celula_soma_2 = janeiro.cell(row=total, column=2, value=soma_coluna_2)
celula_soma_7 = janeiro.cell(row=total, column=7, value=soma_coluna_7)
celula_soma_2.font = Font(bold=True)
celula_soma_7.font = Font(bold=True)
janeiro.cell(row=50, column=1, value="Total").font = Font(bold=True)
janeiro.cell(row=50, column=6, value="Total").font = Font(bold=True)
#pagina 2 (fevereiro)
fevereiro = planilha.create_sheet('Fevereiro')
cabecalho = ["Compra", "Preço", "Data", "Loja", " ", "Compra", "Preço", "Data", "Loja"]
for col, valor in enumerate(cabecalho, start=1):
    celula = fevereiro.cell(row=1, column=col, value=valor)
    celula.font = Font(bold=True)
    celula.fill = cinza
for row in range(2, 51):
    celula_coluna_2 = fevereiro.cell(row=row, column=2)
    celula_coluna_7 = fevereiro.cell(row=row, column=7)
    celula_coluna_2.style = moeda
    celula_coluna_7.style = moeda
for col, valor in enumerate(cabecalho, start=1):
    celula = fevereiro.cell(row=1, column=col, value=valor)
    celula.font = Font(bold=True)
for i in range(1, num_linhas + 1):
    for j in range(1, num_colunas + 1):
        cell = fevereiro.cell(row=i, column=j)
        cell.border = contorno
fevereiro.column_dimensions[openpyxl.utils.get_column_letter(divisoria)].width = 0.5
fevereiro.column_dimensions[openpyxl.utils.get_column_letter(compra1)].width = 30
fevereiro.column_dimensions[openpyxl.utils.get_column_letter(loja1)].width = 30
fevereiro.column_dimensions[openpyxl.utils.get_column_letter(compra2)].width = 30
fevereiro.column_dimensions[openpyxl.utils.get_column_letter(loja2)].width = 30
fevereiro.column_dimensions[openpyxl.utils.get_column_letter(preco1)].width = 11
fevereiro.column_dimensions[openpyxl.utils.get_column_letter(preco2)].width = 11
fevereiro.column_dimensions[openpyxl.utils.get_column_letter(data1)].width = 5
fevereiro.column_dimensions[openpyxl.utils.get_column_letter(data2)].width = 5
for row in range(1, 51):
    celula = fevereiro.cell(row=row, column=divisoria)
    celula.fill = preenchimento
celula_soma_2 = fevereiro.cell(row=total, column=2, value=soma_coluna_2)
celula_soma_7 = fevereiro.cell(row=total, column=7, value=soma_coluna_7)
celula_soma_2.font = Font(bold=True)
celula_soma_7.font = Font(bold=True)
fevereiro.cell(row=50, column=1, value="Total").font = Font(bold=True)
fevereiro.cell(row=50, column=6, value="Total").font = Font(bold=True)
#página 3 (março)
marco = planilha.create_sheet('Março')
cabecalho = ["Compra", "Preço", "Data", "Loja"," ", "Compra", "Preço", "Data", "Loja"]
for col, valor in enumerate(cabecalho, start=1):
    celula = marco.cell(row=1, column=col, value=valor)
    celula.font = Font(bold=True)
    celula.fill = cinza
for row in range(2, 51):
    celula_coluna_2 = marco.cell(row=row, column=2)
    celula_coluna_7 = marco.cell(row=row, column=7)
    celula_coluna_2.style = moeda
    celula_coluna_7.style = moeda
for col, valor in enumerate(cabecalho, start=1):
    celula = marco.cell(row=1, column=col, value=valor)
    celula.font = Font(bold=True)
for i in range(1, num_linhas + 1):
    for j in range(1, num_colunas + 1):
        cell = marco.cell(row=i, column=j)
        cell.border = contorno
marco.column_dimensions[openpyxl.utils.get_column_letter(divisoria)].width = 0.5
marco.column_dimensions[openpyxl.utils.get_column_letter(compra1)].width = 30
marco.column_dimensions[openpyxl.utils.get_column_letter(loja1)].width = 30
marco.column_dimensions[openpyxl.utils.get_column_letter(compra2)].width = 30
marco.column_dimensions[openpyxl.utils.get_column_letter(loja2)].width = 30
marco.column_dimensions[openpyxl.utils.get_column_letter(preco1)].width = 11
marco.column_dimensions[openpyxl.utils.get_column_letter(preco2)].width = 11
marco.column_dimensions[openpyxl.utils.get_column_letter(data1)].width = 5
marco.column_dimensions[openpyxl.utils.get_column_letter(data2)].width = 5
for row in range(1, 51):
    celula = marco.cell(row=row, column=divisoria)
    celula.fill = preenchimento
celula_soma_2 = marco.cell(row=total, column=2, value=soma_coluna_2)
celula_soma_7 = marco.cell(row=total, column=7, value=soma_coluna_7)
celula_soma_2.font = Font(bold=True)
celula_soma_7.font = Font(bold=True)
marco.cell(row=50, column=1, value="Total").font = Font(bold=True)
marco.cell(row=50, column=6, value="Total").font = Font(bold=True)
#página 4 (abril)
abril = planilha.create_sheet('Abril')
cabecalho = ["Compra", "Preço", "Data", "Loja"," ", "Compra", "Preço", "Data", "Loja"]
for col, valor in enumerate(cabecalho, start=1):
    celula = abril.cell(row=1, column=col, value=valor)
    celula.font = Font(bold=True)
    celula.fill = cinza
for row in range(2, 51):
    celula_coluna_2 = abril.cell(row=row, column=2)
    celula_coluna_7 = abril.cell(row=row, column=7)
    celula_coluna_2.style = moeda
    celula_coluna_7.style = moeda
for col, valor in enumerate(cabecalho, start=1):
    celula = abril.cell(row=1, column=col, value=valor)
    celula.font = Font(bold=True)
for i in range(1, num_linhas + 1):
    for j in range(1, num_colunas + 1):
        cell = abril.cell(row=i, column=j)
        cell.border = contorno
abril.column_dimensions[openpyxl.utils.get_column_letter(divisoria)].width = 0.5
abril.column_dimensions[openpyxl.utils.get_column_letter(compra1)].width = 30
abril.column_dimensions[openpyxl.utils.get_column_letter(loja1)].width = 30
abril.column_dimensions[openpyxl.utils.get_column_letter(compra2)].width = 30
abril.column_dimensions[openpyxl.utils.get_column_letter(loja2)].width = 30
abril.column_dimensions[openpyxl.utils.get_column_letter(preco1)].width = 11
abril.column_dimensions[openpyxl.utils.get_column_letter(preco2)].width = 11
abril.column_dimensions[openpyxl.utils.get_column_letter(data1)].width = 5
abril.column_dimensions[openpyxl.utils.get_column_letter(data2)].width = 5
for row in range(1, 51):
    celula = abril.cell(row=row, column=divisoria)
    celula.fill = preenchimento
celula_soma_2 = abril.cell(row=total, column=2, value=soma_coluna_2)
celula_soma_7 = abril.cell(row=total, column=7, value=soma_coluna_7)
celula_soma_2.font = Font(bold=True)
celula_soma_7.font = Font(bold=True)
abril.cell(row=50, column=1, value="Total").font = Font(bold=True)
abril.cell(row=50, column=6, value="Total").font = Font(bold=True)
#página 5 (maio)
maio = planilha.create_sheet('Maio')
cabecalho = ["Compra", "Preço", "Data", "Loja"," ", "Compra", "Preço", "Data", "Loja"]
for col, valor in enumerate(cabecalho, start=1):
    celula = maio.cell(row=1, column=col, value=valor)
    celula.font = Font(bold=True)
    celula.fill = cinza
for row in range(2, 51):
    celula_coluna_2 = maio.cell(row=row, column=2)
    celula_coluna_7 = maio.cell(row=row, column=7)
    celula_coluna_2.style = moeda
    celula_coluna_7.style = moeda
for col, valor in enumerate(cabecalho, start=1):
    celula = maio.cell(row=1, column=col, value=valor)
    celula.font = Font(bold=True)
for i in range(1, num_linhas + 1):
    for j in range(1, num_colunas + 1):
        cell = maio.cell(row=i, column=j)
        cell.border = contorno
maio.column_dimensions[openpyxl.utils.get_column_letter(divisoria)].width = 0.5
maio.column_dimensions[openpyxl.utils.get_column_letter(compra1)].width = 30
maio.column_dimensions[openpyxl.utils.get_column_letter(loja1)].width = 30
maio.column_dimensions[openpyxl.utils.get_column_letter(compra2)].width = 30
maio.column_dimensions[openpyxl.utils.get_column_letter(loja2)].width = 30
maio.column_dimensions[openpyxl.utils.get_column_letter(preco1)].width = 11
maio.column_dimensions[openpyxl.utils.get_column_letter(preco2)].width = 11
maio.column_dimensions[openpyxl.utils.get_column_letter(data1)].width = 5
maio.column_dimensions[openpyxl.utils.get_column_letter(data2)].width = 5
for row in range(1, 51):
    celula = maio.cell(row=row, column=divisoria)
    celula.fill = preenchimento
celula_soma_2 = maio.cell(row=total, column=2, value=soma_coluna_2)
celula_soma_7 = maio.cell(row=total, column=7, value=soma_coluna_7)
celula_soma_2.font = Font(bold=True)
celula_soma_7.font = Font(bold=True)
maio.cell(row=50, column=1, value="Total").font = Font(bold=True)
maio.cell(row=50, column=6, value="Total").font = Font(bold=True)
#página 6 (junho)
junho = planilha.create_sheet('Junho')
cabecalho = ["Compra", "Preço", "Data", "Loja"," ", "Compra", "Preço", "Data", "Loja"]
for col, valor in enumerate(cabecalho, start=1):
    celula = junho.cell(row=1, column=col, value=valor)
    celula.font = Font(bold=True)
    celula.fill = cinza
for row in range(2, 51):
    celula_coluna_2 = junho.cell(row=row, column=2)
    celula_coluna_7 = junho.cell(row=row, column=7)
    celula_coluna_2.style = moeda
    celula_coluna_7.style = moeda
for col, valor in enumerate(cabecalho, start=1):
    celula = junho.cell(row=1, column=col, value=valor)
    celula.font = Font(bold=True)
for i in range(1, num_linhas + 1):
    for j in range(1, num_colunas + 1):
        cell = junho.cell(row=i, column=j)
        cell.border = contorno
junho.column_dimensions[openpyxl.utils.get_column_letter(divisoria)].width = 0.5
junho.column_dimensions[openpyxl.utils.get_column_letter(compra1)].width = 30
junho.column_dimensions[openpyxl.utils.get_column_letter(loja1)].width = 30
junho.column_dimensions[openpyxl.utils.get_column_letter(compra2)].width = 30
junho.column_dimensions[openpyxl.utils.get_column_letter(loja2)].width = 30
junho.column_dimensions[openpyxl.utils.get_column_letter(preco1)].width = 11
junho.column_dimensions[openpyxl.utils.get_column_letter(preco2)].width = 11
junho.column_dimensions[openpyxl.utils.get_column_letter(data1)].width = 5
junho.column_dimensions[openpyxl.utils.get_column_letter(data2)].width = 5
for row in range(1, 51):
    celula = junho.cell(row=row, column=divisoria)
    celula.fill = preenchimento
celula_soma_2 = junho.cell(row=total, column=2, value=soma_coluna_2)
celula_soma_7 = junho.cell(row=total, column=7, value=soma_coluna_7)
celula_soma_2.font = Font(bold=True)
celula_soma_7.font = Font(bold=True)
junho.cell(row=50, column=1, value="Total").font = Font(bold=True)
junho.cell(row=50, column=6, value="Total").font = Font(bold=True)
#página 7 (julho)
julho = planilha.create_sheet('Julho')
cabecalho = ["Compra", "Preço", "Data", "Loja"," ", "Compra", "Preço", "Data", "Loja"]
for col, valor in enumerate(cabecalho, start=1):
    celula = julho.cell(row=1, column=col, value=valor)
    celula.font = Font(bold=True)
    celula.fill = cinza
for row in range(2, 51):
    celula_coluna_2 = julho.cell(row=row, column=2)
    celula_coluna_7 = julho.cell(row=row, column=7)
    celula_coluna_2.style = moeda
    celula_coluna_7.style = moeda
for col, valor in enumerate(cabecalho, start=1):
    celula = julho.cell(row=1, column=col, value=valor)
    celula.font = Font(bold=True)
for i in range(1, num_linhas + 1):
    for j in range(1, num_colunas + 1):
        cell = julho.cell(row=i, column=j)
        cell.border = contorno
julho.column_dimensions[openpyxl.utils.get_column_letter(divisoria)].width = 0.5
julho.column_dimensions[openpyxl.utils.get_column_letter(compra1)].width = 30
julho.column_dimensions[openpyxl.utils.get_column_letter(loja1)].width = 30
julho.column_dimensions[openpyxl.utils.get_column_letter(compra2)].width = 30
julho.column_dimensions[openpyxl.utils.get_column_letter(loja2)].width = 30
julho.column_dimensions[openpyxl.utils.get_column_letter(preco1)].width = 11
julho.column_dimensions[openpyxl.utils.get_column_letter(preco2)].width = 11
julho.column_dimensions[openpyxl.utils.get_column_letter(data1)].width = 5
julho.column_dimensions[openpyxl.utils.get_column_letter(data2)].width = 5
for row in range(1, 51):
    celula = julho.cell(row=row, column=divisoria)
    celula.fill = preenchimento
celula_soma_2 = julho.cell(row=total, column=2, value=soma_coluna_2)
celula_soma_7 = julho.cell(row=total, column=7, value=soma_coluna_7)
celula_soma_2.font = Font(bold=True)
celula_soma_7.font = Font(bold=True)
julho.cell(row=50, column=1, value="Total").font = Font(bold=True)
julho.cell(row=50, column=6, value="Total").font = Font(bold=True)
#página 8 (agosto)
agosto = planilha.create_sheet('Agosto')
cabecalho = ["Compra", "Preço", "Data", "Loja"," ", "Compra", "Preço", "Data", "Loja"]
for col, valor in enumerate(cabecalho, start=1):
    celula = agosto.cell(row=1, column=col, value=valor)
    celula.font = Font(bold=True)
    celula.fill = cinza
for row in range(2, 51):
    celula_coluna_2 = agosto.cell(row=row, column=2)
    celula_coluna_7 = agosto.cell(row=row, column=7)
    celula_coluna_2.style = moeda
    celula_coluna_7.style = moeda
for col, valor in enumerate(cabecalho, start=1):
    celula = agosto.cell(row=1, column=col, value=valor)
    celula.font = Font(bold=True)
for i in range(1, num_linhas + 1):
    for j in range(1, num_colunas + 1):
        cell = agosto.cell(row=i, column=j)
        cell.border = contorno
agosto.column_dimensions[openpyxl.utils.get_column_letter(divisoria)].width = 0.5
agosto.column_dimensions[openpyxl.utils.get_column_letter(compra1)].width = 30
agosto.column_dimensions[openpyxl.utils.get_column_letter(loja1)].width = 30
agosto.column_dimensions[openpyxl.utils.get_column_letter(compra2)].width = 30
agosto.column_dimensions[openpyxl.utils.get_column_letter(loja2)].width = 30
agosto.column_dimensions[openpyxl.utils.get_column_letter(preco1)].width = 11
agosto.column_dimensions[openpyxl.utils.get_column_letter(preco2)].width = 11
agosto.column_dimensions[openpyxl.utils.get_column_letter(data1)].width = 5
agosto.column_dimensions[openpyxl.utils.get_column_letter(data2)].width = 5
for row in range(1, 51):
    celula = agosto.cell(row=row, column=divisoria)
    celula.fill = preenchimento
celula_soma_2 = agosto.cell(row=total, column=2, value=soma_coluna_2)
celula_soma_7 = agosto.cell(row=total, column=7, value=soma_coluna_7)
celula_soma_2.font = Font(bold=True)
celula_soma_7.font = Font(bold=True)
agosto.cell(row=50, column=1, value="Total").font = Font(bold=True)
agosto.cell(row=50, column=6, value="Total").font = Font(bold=True)
#página 9 (setembro)
setembro = planilha.create_sheet('Setembro')
cabecalho = ["Compra", "Preço", "Data", "Loja"," ", "Compra", "Preço", "Data", "Loja"]
for col, valor in enumerate(cabecalho, start=1):
    celula = setembro.cell(row=1, column=col, value=valor)
    celula.font = Font(bold=True)
    celula.fill = cinza
for row in range(2, 51):
    celula_coluna_2 = setembro.cell(row=row, column=2)
    celula_coluna_7 = setembro.cell(row=row, column=7)
    celula_coluna_2.style = moeda
    celula_coluna_7.style = moeda
for col, valor in enumerate(cabecalho, start=1):
    celula = setembro.cell(row=1, column=col, value=valor)
    celula.font = Font(bold=True)
for i in range(1, num_linhas + 1):
    for j in range(1, num_colunas + 1):
        cell = setembro.cell(row=i, column=j)
        cell.border = contorno
setembro.column_dimensions[openpyxl.utils.get_column_letter(divisoria)].width = 0.5
setembro.column_dimensions[openpyxl.utils.get_column_letter(compra1)].width = 30
setembro.column_dimensions[openpyxl.utils.get_column_letter(loja1)].width = 30
setembro.column_dimensions[openpyxl.utils.get_column_letter(compra2)].width = 30
setembro.column_dimensions[openpyxl.utils.get_column_letter(loja2)].width = 30
setembro.column_dimensions[openpyxl.utils.get_column_letter(preco1)].width = 11
setembro.column_dimensions[openpyxl.utils.get_column_letter(preco2)].width = 11
setembro.column_dimensions[openpyxl.utils.get_column_letter(data1)].width = 5
setembro.column_dimensions[openpyxl.utils.get_column_letter(data2)].width = 5
for row in range(1, 51):
    celula = setembro.cell(row=row, column=divisoria)
    celula.fill = preenchimento
celula_soma_2 = setembro.cell(row=total, column=2, value=soma_coluna_2)
celula_soma_7 = setembro.cell(row=total, column=7, value=soma_coluna_7)
celula_soma_2.font = Font(bold=True)
celula_soma_7.font = Font(bold=True)
setembro.cell(row=50, column=1, value="Total").font = Font(bold=True)
setembro.cell(row=50, column=6, value="Total").font = Font(bold=True)
#página 10 (outubro)
outubro = planilha.create_sheet('Outubro')
cabecalho = ["Compra", "Preço", "Data", "Loja"," ", "Compra", "Preço", "Data", "Loja"]
for col, valor in enumerate(cabecalho, start=1):
    celula = outubro.cell(row=1, column=col, value=valor)
    celula.font = Font(bold=True)
    celula.fill = cinza
for row in range(2, 51):
    celula_coluna_2 = outubro.cell(row=row, column=2)
    celula_coluna_7 = outubro.cell(row=row, column=7)
    celula_coluna_2.style = moeda
    celula_coluna_7.style = moeda
for col, valor in enumerate(cabecalho, start=1):
    celula = outubro.cell(row=1, column=col, value=valor)
    celula.font = Font(bold=True)
for i in range(1, num_linhas + 1):
    for j in range(1, num_colunas + 1):
        cell = outubro.cell(row=i, column=j)
        cell.border = contorno
outubro.column_dimensions[openpyxl.utils.get_column_letter(divisoria)].width = 0.5
outubro.column_dimensions[openpyxl.utils.get_column_letter(compra1)].width = 30
outubro.column_dimensions[openpyxl.utils.get_column_letter(loja1)].width = 30
outubro.column_dimensions[openpyxl.utils.get_column_letter(compra2)].width = 30
outubro.column_dimensions[openpyxl.utils.get_column_letter(loja2)].width = 30
outubro.column_dimensions[openpyxl.utils.get_column_letter(preco1)].width = 11
outubro.column_dimensions[openpyxl.utils.get_column_letter(preco2)].width = 11
outubro.column_dimensions[openpyxl.utils.get_column_letter(data1)].width = 5
outubro.column_dimensions[openpyxl.utils.get_column_letter(data2)].width = 5
for row in range(1, 51):
    celula = outubro.cell(row=row, column=divisoria)
    celula.fill = preenchimento
celula_soma_2 = outubro.cell(row=total, column=2, value=soma_coluna_2)
celula_soma_7 = outubro.cell(row=total, column=7, value=soma_coluna_7)
celula_soma_2.font = Font(bold=True)
celula_soma_7.font = Font(bold=True)
outubro.cell(row=50, column=1, value="Total").font = Font(bold=True)
outubro.cell(row=50, column=6, value="Total").font = Font(bold=True)
#página 11 (novembro)
novembro = planilha.create_sheet('Novembro')
cabecalho = ["Compra", "Preço", "Data", "Loja"," ", "Compra", "Preço", "Data", "Loja"]
for col, valor in enumerate(cabecalho, start=1):
    celula = novembro.cell(row=1, column=col, value=valor)
    celula.font = Font(bold=True)
    celula.fill = cinza
for row in range(2, 51):
    celula_coluna_2 = novembro.cell(row=row, column=2)
    celula_coluna_7 = novembro.cell(row=row, column=7)
    celula_coluna_2.style = moeda
    celula_coluna_7.style = moeda
for col, valor in enumerate(cabecalho, start=1):
    celula = novembro.cell(row=1, column=col, value=valor)
    celula.font = Font(bold=True)
for i in range(1, num_linhas + 1):
    for j in range(1, num_colunas + 1):
        cell = novembro.cell(row=i, column=j)
        cell.border = contorno
novembro.column_dimensions[openpyxl.utils.get_column_letter(divisoria)].width = 0.5
novembro.column_dimensions[openpyxl.utils.get_column_letter(compra1)].width = 30
novembro.column_dimensions[openpyxl.utils.get_column_letter(loja1)].width = 30
novembro.column_dimensions[openpyxl.utils.get_column_letter(compra2)].width = 30
novembro.column_dimensions[openpyxl.utils.get_column_letter(loja2)].width = 30
novembro.column_dimensions[openpyxl.utils.get_column_letter(preco1)].width = 11
novembro.column_dimensions[openpyxl.utils.get_column_letter(preco2)].width = 11
novembro.column_dimensions[openpyxl.utils.get_column_letter(data1)].width = 5
novembro.column_dimensions[openpyxl.utils.get_column_letter(data2)].width = 5
for row in range(1, 51):
    celula = novembro.cell(row=row, column=divisoria)
    celula.fill = preenchimento
celula_soma_2 = novembro.cell(row=total, column=2, value=soma_coluna_2)
celula_soma_7 = novembro.cell(row=total, column=7, value=soma_coluna_7)
celula_soma_2.font = Font(bold=True)
celula_soma_7.font = Font(bold=True)
novembro.cell(row=50, column=1, value="Total").font = Font(bold=True)
novembro.cell(row=50, column=6, value="Total").font = Font(bold=True)
#página 12 (dezembro)
dezembro = planilha.create_sheet('Dezembro')
cabecalho = ["Compra", "Preço", "Data", "Loja"," ", "Compra", "Preço", "Data", "Loja"]
for col, valor in enumerate(cabecalho, start=1):
    celula = dezembro.cell(row=1, column=col, value=valor)
    celula.font = Font(bold=True)
    celula.fill = cinza
for row in range(2, 51):
    celula_coluna_2 = dezembro.cell(row=row, column=2)
    celula_coluna_7 = dezembro.cell(row=row, column=7)
    celula_coluna_2.style = moeda
    celula_coluna_7.style = moeda
for col, valor in enumerate(cabecalho, start=1):
    celula = dezembro.cell(row=1, column=col, value=valor)
    celula.font = Font(bold=True)
for i in range(1, num_linhas + 1):
    for j in range(1, num_colunas + 1):
        cell = dezembro.cell(row=i, column=j)
        cell.border = contorno
dezembro.column_dimensions[openpyxl.utils.get_column_letter(divisoria)].width = 0.5
dezembro.column_dimensions[openpyxl.utils.get_column_letter(compra1)].width = 30
dezembro.column_dimensions[openpyxl.utils.get_column_letter(loja1)].width = 30
dezembro.column_dimensions[openpyxl.utils.get_column_letter(compra2)].width = 30
dezembro.column_dimensions[openpyxl.utils.get_column_letter(loja2)].width = 30
dezembro.column_dimensions[openpyxl.utils.get_column_letter(preco1)].width = 11
dezembro.column_dimensions[openpyxl.utils.get_column_letter(preco2)].width = 11
dezembro.column_dimensions[openpyxl.utils.get_column_letter(data1)].width = 5
dezembro.column_dimensions[openpyxl.utils.get_column_letter(data2)].width = 5
for row in range(1, 51):
    celula = dezembro.cell(row=row, column=divisoria)
    celula.fill = preenchimento
celula_soma_2 = dezembro.cell(row=total, column=2, value=soma_coluna_2)
celula_soma_7 = dezembro.cell(row=total, column=7, value=soma_coluna_7)
celula_soma_2.font = Font(bold=True)
celula_soma_7.font = Font(bold=True)
dezembro.cell(row=50, column=1, value="Total").font = Font(bold=True)
dezembro.cell(row=50, column=6, value="Total").font = Font(bold=True)
#salvar planilha
planilha.save("panilha_gastos.xlsx")
print("Planilha de gastos criada com sucesso.")
