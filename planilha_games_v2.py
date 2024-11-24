import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import tkinter as tk
from tkinter import filedialog
import pandas as pd
planilha = openpyxl.Workbook()
sheet = planilha.active
planilha.remove(sheet)
#definir bordas
contorno = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)
#definir colunas
numero = 1
game = 2
plataforma = 3
mes = 4
nota = 5
#definir tamanho da planilha
num_linhas = 40
num_colunas = 5
media_notas = f'=AVERAGE(E3:E{40})' #definir média
cinza = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid") #definir preenchimento
zerados = planilha.create_sheet('Games zerados') #criar página
cabecalho = ['Nº', 'Game', 'Plataforma', 'Mês', 'Nota'] #definir cabeçalho
for col, valor in enumerate(cabecalho, start=1):
    celula = zerados.cell(row=1, column=col, value=valor)
    celula.font = Font(bold=True)
    celula.fill = cinza
for col in range(1, 5):
    celula = zerados.cell(row=2, column=col)
    celula.font = Font(bold=True)
    celula.fill = cinza
zerados.merge_cells('A2:D2')
zerados['A2'] = 'Média:   '
zerados['A2'].alignment = Alignment(horizontal='left')
zerados['E2'].fill = cinza
zerados['E2'].font = Font(bold=True)
zerados.freeze_panes= "A3"
#aplicar bordas
for i in range(1, num_linhas + 1):
    for j in range(1, num_colunas + 1):
        cell = zerados.cell(row=i, column=j)
        cell.border = contorno
#formatar colunas
zerados.column_dimensions[openpyxl.utils.get_column_letter(numero)].width = 3
zerados.column_dimensions[openpyxl.utils.get_column_letter(game)].width = 40
zerados.column_dimensions[openpyxl.utils.get_column_letter(plataforma)].width = 15
zerados.column_dimensions[openpyxl.utils.get_column_letter(mes)].width = 10
zerados.column_dimensions[openpyxl.utils.get_column_letter(nota)].width = 5
celula_media = zerados.cell(row=2, column=5, value=media_notas)
centralizar = Alignment(horizontal='center')
#salvar planilha
def salvar_planilha():
    root = tk.Tk()
    root.withdraw()
    arquivo = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")],
        title="Salvar planilha como"
    )
    if arquivo:
        planilha.save(arquivo)
        print(f"Planilha salva em: {arquivo}")
    else:
        print("Nenhum arquivo foi selecionado.")
if __name__ == "__main__":
    salvar_planilha()

