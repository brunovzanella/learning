import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

planilha = openpyxl.Workbook()
sheet = planilha.active
planilha.remove(sheet)
#definir bordas
contorno = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)
#definir colunas
numero = 1
game = 2
plataforma = 3
mes = 4
nota = 5
#definir tamanho da planilha
num_linhas = 40
num_colunas = 5
media_notas = f'=AVERAGE(E3:E{40})' #definir média
cinza = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid") #definir preenchimento
games2025 = planilha.create_sheet('2025') #criar página
cabecalho = ['Nº', 'Game', 'Plataforma', 'Mês', 'Nota'] #definir cabeçalho
for col, valor in enumerate(cabecalho, start=1):
    celula = games2025.cell(row=1, column=col, value=valor)
    celula.font = Font(bold=True)
    celula.fill = cinza
for col in range(1, 5):
    celula = games2025.cell(row=2, column=col)
    celula.font = Font(bold=True)
    celula.fill = cinza
games2025.merge_cells('A2:D2')
games2025['A2'] = 'Média:   '
games2025['A2'].alignment = Alignment(horizontal='right')
games2025['E2'].fill = cinza
games2025['E2'].font = Font(bold=True)
games2025.freeze_panes= "A3"
#aplicar bordas
for i in range(1, num_linhas + 1):
    for j in range(1, num_colunas + 1):
        cell = games2025.cell(row=i, column=j)
        cell.border = contorno
#formatar colunas
games2025.column_dimensions[openpyxl.utils.get_column_letter(numero)].width = 5
games2025.column_dimensions[openpyxl.utils.get_column_letter(game)].width = 40
games2025.column_dimensions[openpyxl.utils.get_column_letter(plataforma)].width = 20
games2025.column_dimensions[openpyxl.utils.get_column_letter(mes)].width = 15
games2025.column_dimensions[openpyxl.utils.get_column_letter(nota)].width = 5
celula_media = games2025.cell(row=2, column=5, value=media_notas)
games2025.cell(row=2, column=1)
#salvar planilha
planilha.save(r"C:\Users\bruno\OneDrive\Documents\Games.xlsx")
print("Planilha de games criada com sucesso.")

