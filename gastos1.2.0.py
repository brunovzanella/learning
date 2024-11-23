'''
versão 1.1.0
patch notes:
* linha 1 agora categoriza gastos
* criada terceira categoria de gastos para separar vale alimentação e mobilidade
* cabeçalho passou para a linha 2
versão 1.2.0
* passado o resultado da soma para a linha 3
* colorida a linha 3 de cinza
* adicionada função linhas de congelamento nas linhas 1, 2 e 3
'''
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, NamedStyle, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.protection import SheetProtection
#criar arquivo xlxs
planilha = openpyxl.Workbook()
#excluir página padrão
sheet = planilha.active
planilha.remove(sheet)
#definir bordas
contorno = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)
#definir colunas
compra1 = 1
preco1 = 2
data1 = 3
loja1 = 4
divisoria = 5
compra2 = 6
preco2 = 7
data2 = 8
loja2 = 9
divisoria2 = 10
compra3 = 11
preco3 = 12
data3 = 13
loja3 = 14
#definir tamanho da tabela
num_linhas = 50
num_colunas = 14
#definir soma
total = 3 
soma_coluna_2 = f'=SUM(B4:B{50})'
soma_coluna_7 = f'=SUM(G4:G{50})'
soma_coluna_12 = f'=SUM(L4:L{50})'
#definir preenchimentos 
preto = PatternFill(start_color="FF000000", end_color="FF000000", fill_type="solid")
cinza = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
#definir formato moeda
moeda = NamedStyle(name="moeda")
moeda.number_format = 'R$ #,##0.00'
#criar páginas com linhas e colunas formatadas
janeiro = planilha.create_sheet('Janeiro')
janeiro.merge_cells('A1:D1')
janeiro['A1'] = 'Cartão de crédito'
janeiro['A1'].alignment = Alignment(horizontal='center', vertical='center')
janeiro['A1'].font = Font(color="FFFFFF", bold=True, size=14)
janeiro['A1'].fill = preto
janeiro.merge_cells('F1:I1')
janeiro['F1'] = 'Vale alimentação'
janeiro['F1'].alignment = Alignment(horizontal='center', vertical='center')
janeiro['F1'].font = Font(color="FFFFFF", bold=True, size=14)
janeiro['f1'].fill = preto
janeiro.merge_cells('K1:N1')
janeiro['K1'] = 'Vale mobilidade'
janeiro['K1'].alignment = Alignment(horizontal='center', vertical='center')
janeiro['K1'].font = Font(color="FFFFFF", bold=True, size=14)
janeiro['K1'].fill = preto
cabecalho = ["Compra", "Preço", "Data", "Loja", " ", "Compra", "Preço", "Data", "Loja", " ", "Compra", "Preço", "Data", "Loja"]
for col, valor in enumerate(cabecalho, start=1):
    celula = janeiro.cell(row=2, column=col, value=valor)
    celula.font = Font(bold=True)
    celula.fill = cinza
for col in range(1, 15):
    celula = janeiro.cell(row=3, column=col)
    celula.fill = cinza
janeiro.freeze_panes = "A4"
#aplicar estilo moeda
for row in range(3, 51):
    celula_coluna_2 = janeiro.cell(row=row, column=2)
    celula_coluna_7 = janeiro.cell(row=row, column=7)
    celula_coluna_12 = janeiro.cell(row=row, column=12)
    celula_coluna_2.style = moeda
    celula_coluna_7.style = moeda
    celula_coluna_12.style = moeda
#aplicar bordas
for i in range(1, num_linhas + 1):
    for j in range(1, num_colunas + 1):
        cell = janeiro.cell(row=i, column=j)
        cell.border = contorno
#formatar colunas
janeiro.column_dimensions[openpyxl.utils.get_column_letter(divisoria)].width = 0.5
janeiro.column_dimensions[openpyxl.utils.get_column_letter(divisoria2)].width = 0.5
janeiro.column_dimensions[openpyxl.utils.get_column_letter(compra1)].width = 30
janeiro.column_dimensions[openpyxl.utils.get_column_letter(compra2)].width = 30
janeiro.column_dimensions[openpyxl.utils.get_column_letter(compra3)].width = 30
janeiro.column_dimensions[openpyxl.utils.get_column_letter(loja1)].width = 30
janeiro.column_dimensions[openpyxl.utils.get_column_letter(loja2)].width = 30
janeiro.column_dimensions[openpyxl.utils.get_column_letter(loja3)].width = 30
janeiro.column_dimensions[openpyxl.utils.get_column_letter(preco1)].width = 11
janeiro.column_dimensions[openpyxl.utils.get_column_letter(preco2)].width = 11
janeiro.column_dimensions[openpyxl.utils.get_column_letter(preco3)].width = 11
janeiro.column_dimensions[openpyxl.utils.get_column_letter(data1)].width = 5
janeiro.column_dimensions[openpyxl.utils.get_column_letter(data2)].width = 5
janeiro.column_dimensions[openpyxl.utils.get_column_letter(data3)].width = 5
for row in range(1, 51):
    celula = janeiro.cell(row=row, column=divisoria)
    celula.fill = preto
for row in range(1, 51):
    celula = janeiro.cell(row=row, column=divisoria2)
    celula.fill = preto
celula_soma_2 = janeiro.cell(row=3, column=2, value=soma_coluna_2)
celula_soma_7 = janeiro.cell(row=3, column=7, value=soma_coluna_7)
celula_soma_12 = janeiro.cell(row=3, column=12, value=soma_coluna_12)
celula_soma_2.font = Font(bold=True)
celula_soma_2.fill = cinza
celula_soma_7.font = Font(bold=True)
celula_soma_7.fill = cinza
celula_soma_12.font = Font(bold=True)
celula_soma_12.fill = cinza
janeiro.cell(row=3, column=1, value="Total").font = Font(bold=True)
janeiro.cell(row=3, column=6, value="Total").font = Font(bold=True)
janeiro.cell(row=3, column=11, value="Total").font =Font(bold=True)
#pagina 2 (fevereiro)
fevereiro = planilha.create_sheet('Fevereiro')
fevereiro.merge_cells('A1:D1')
fevereiro['A1'] = 'Cartão de crédito'
fevereiro['A1'].alignment = Alignment(horizontal='center', vertical='center')
fevereiro['A1'].font = Font(color="FFFFFF", bold=True, size=14)
fevereiro['A1'].fill = preto
fevereiro.merge_cells('F1:I1')
fevereiro['F1'] = 'Vale alimentação'
fevereiro['F1'].alignment = Alignment(horizontal='center', vertical='center')
fevereiro['F1'].font = Font(color="FFFFFF", bold=True, size=14)
fevereiro['f1'].fill = preto
fevereiro.merge_cells('K1:N1')
fevereiro['K1'] = 'Vale mobilidade'
fevereiro['K1'].alignment = Alignment(horizontal='center', vertical='center')
fevereiro['K1'].font = Font(color="FFFFFF", bold=True, size=14)
fevereiro['K1'].fill = preto
cabecalho = ["Compra", "Preço", "Data", "Loja", " ", "Compra", "Preço", "Data", "Loja", " ", "Compra", "Preço", "Data", "Loja"]
for col, valor in enumerate(cabecalho, start=1):
    celula = fevereiro.cell(row=2, column=col, value=valor)
    celula.font = Font(bold=True)
    celula.fill = cinza
for col in range(1, 15):
    celula = fevereiro.cell(row=3, column=col)
    celula.fill = cinza
fevereiro.freeze_panes = "A4"
#aplicar estilo moeda
for row in range(3, 51):
    celula_coluna_2 = fevereiro.cell(row=row, column=2)
    celula_coluna_7 = fevereiro.cell(row=row, column=7)
    celula_coluna_12 = fevereiro.cell(row=row, column=12)
    celula_coluna_2.style = moeda
    celula_coluna_7.style = moeda
    celula_coluna_12.style = moeda
#aplicar bordas
for i in range(1, num_linhas + 1):
    for j in range(1, num_colunas + 1):
        cell = fevereiro.cell(row=i, column=j)
        cell.border = contorno
#formatar colunas
fevereiro.column_dimensions[openpyxl.utils.get_column_letter(divisoria)].width = 0.5
fevereiro.column_dimensions[openpyxl.utils.get_column_letter(divisoria2)].width = 0.5
fevereiro.column_dimensions[openpyxl.utils.get_column_letter(compra1)].width = 30
fevereiro.column_dimensions[openpyxl.utils.get_column_letter(compra2)].width = 30
fevereiro.column_dimensions[openpyxl.utils.get_column_letter(compra3)].width = 30
fevereiro.column_dimensions[openpyxl.utils.get_column_letter(loja1)].width = 30
fevereiro.column_dimensions[openpyxl.utils.get_column_letter(loja2)].width = 30
fevereiro.column_dimensions[openpyxl.utils.get_column_letter(loja3)].width = 30
fevereiro.column_dimensions[openpyxl.utils.get_column_letter(preco1)].width = 11
fevereiro.column_dimensions[openpyxl.utils.get_column_letter(preco2)].width = 11
fevereiro.column_dimensions[openpyxl.utils.get_column_letter(preco3)].width = 11
fevereiro.column_dimensions[openpyxl.utils.get_column_letter(data1)].width = 5
fevereiro.column_dimensions[openpyxl.utils.get_column_letter(data2)].width = 5
fevereiro.column_dimensions[openpyxl.utils.get_column_letter(data3)].width = 5
for row in range(1, 51):
    celula = fevereiro.cell(row=row, column=divisoria)
    celula.fill = preto
for row in range(1, 51):
    celula = fevereiro.cell(row=row, column=divisoria2)
    celula.fill = preto
celula_soma_2 = fevereiro.cell(row=3, column=2, value=soma_coluna_2)
celula_soma_7 = fevereiro.cell(row=3, column=7, value=soma_coluna_7)
celula_soma_12 = fevereiro.cell(row=3, column=12, value=soma_coluna_12)
celula_soma_2.font = Font(bold=True)
celula_soma_2.fill = cinza
celula_soma_7.font = Font(bold=True)
celula_soma_7.fill = cinza
celula_soma_12.font = Font(bold=True)
celula_soma_12.fill = cinza
fevereiro.cell(row=3, column=1, value="Total").font = Font(bold=True)
fevereiro.cell(row=3, column=6, value="Total").font = Font(bold=True)
fevereiro.cell(row=3, column=11, value="Total").font =Font(bold=True)
#página 3 (março)
marco = planilha.create_sheet('Março')
marco.merge_cells('A1:D1')
marco['A1'] = 'Cartão de crédito'
marco['A1'].alignment = Alignment(horizontal='center', vertical='center')
marco['A1'].font = Font(color="FFFFFF", bold=True, size=14)
marco['A1'].fill = preto
marco.merge_cells('F1:I1')
marco['F1'] = 'Vale alimentação'
marco['F1'].alignment = Alignment(horizontal='center', vertical='center')
marco['F1'].font = Font(color="FFFFFF", bold=True, size=14)
marco['f1'].fill = preto
marco.merge_cells('K1:N1')
marco['K1'] = 'Vale mobilidade'
marco['K1'].alignment = Alignment(horizontal='center', vertical='center')
marco['K1'].font = Font(color="FFFFFF", bold=True, size=14)
marco['K1'].fill = preto
cabecalho = ["Compra", "Preço", "Data", "Loja", " ", "Compra", "Preço", "Data", "Loja", " ", "Compra", "Preço", "Data", "Loja"]
for col, valor in enumerate(cabecalho, start=1):
    celula = marco.cell(row=2, column=col, value=valor)
    celula.font = Font(bold=True)
    celula.fill = cinza
for col in range(1, 15):
    celula = marco.cell(row=3, column=col)
    celula.fill = cinza
marco.freeze_panes = "A4"
#aplicar estilo moeda
for row in range(3, 51):
    celula_coluna_2 = marco.cell(row=row, column=2)
    celula_coluna_7 = marco.cell(row=row, column=7)
    celula_coluna_12 = marco.cell(row=row, column=12)
    celula_coluna_2.style = moeda
    celula_coluna_7.style = moeda
    celula_coluna_12.style = moeda
#aplicar bordas
for i in range(1, num_linhas + 1):
    for j in range(1, num_colunas + 1):
        cell = marco.cell(row=i, column=j)
        cell.border = contorno
#formatar colunas
marco.column_dimensions[openpyxl.utils.get_column_letter(divisoria)].width = 0.5
marco.column_dimensions[openpyxl.utils.get_column_letter(divisoria2)].width = 0.5
marco.column_dimensions[openpyxl.utils.get_column_letter(compra1)].width = 30
marco.column_dimensions[openpyxl.utils.get_column_letter(compra2)].width = 30
marco.column_dimensions[openpyxl.utils.get_column_letter(compra3)].width = 30
marco.column_dimensions[openpyxl.utils.get_column_letter(loja1)].width = 30
marco.column_dimensions[openpyxl.utils.get_column_letter(loja2)].width = 30
marco.column_dimensions[openpyxl.utils.get_column_letter(loja3)].width = 30
marco.column_dimensions[openpyxl.utils.get_column_letter(preco1)].width = 11
marco.column_dimensions[openpyxl.utils.get_column_letter(preco2)].width = 11
marco.column_dimensions[openpyxl.utils.get_column_letter(preco3)].width = 11
marco.column_dimensions[openpyxl.utils.get_column_letter(data1)].width = 5
marco.column_dimensions[openpyxl.utils.get_column_letter(data2)].width = 5
marco.column_dimensions[openpyxl.utils.get_column_letter(data3)].width = 5
for row in range(1, 51):
    celula = marco.cell(row=row, column=divisoria)
    celula.fill = preto
for row in range(1, 51):
    celula = marco.cell(row=row, column=divisoria2)
    celula.fill = preto
celula_soma_2 = marco.cell(row=3, column=2, value=soma_coluna_2)
celula_soma_7 = marco.cell(row=3, column=7, value=soma_coluna_7)
celula_soma_12 = marco.cell(row=3, column=12, value=soma_coluna_12)
celula_soma_2.font = Font(bold=True)
celula_soma_2.fill = cinza
celula_soma_7.font = Font(bold=True)
celula_soma_7.fill = cinza
celula_soma_12.font = Font(bold=True)
celula_soma_12.fill = cinza
marco.cell(row=3, column=1, value="Total").font = Font(bold=True)
marco.cell(row=3, column=6, value="Total").font = Font(bold=True)
marco.cell(row=3, column=11, value="Total").font =Font(bold=True)
#página 4 (abril)
abril = planilha.create_sheet('Abril')
abril.merge_cells('A1:D1')
abril['A1'] = 'Cartão de crédito'
abril['A1'].alignment = Alignment(horizontal='center', vertical='center')
abril['A1'].font = Font(color="FFFFFF", bold=True, size=14)
abril['A1'].fill = preto
abril.merge_cells('F1:I1')
abril['F1'] = 'Vale alimentação'
abril['F1'].alignment = Alignment(horizontal='center', vertical='center')
abril['F1'].font = Font(color="FFFFFF", bold=True, size=14)
abril['f1'].fill = preto
abril.merge_cells('K1:N1')
abril['K1'] = 'Vale mobilidade'
abril['K1'].alignment = Alignment(horizontal='center', vertical='center')
abril['K1'].font = Font(color="FFFFFF", bold=True, size=14)
abril['K1'].fill = preto
cabecalho = ["Compra", "Preço", "Data", "Loja", " ", "Compra", "Preço", "Data", "Loja", " ", "Compra", "Preço", "Data", "Loja"]
for col, valor in enumerate(cabecalho, start=1):
    celula = abril.cell(row=2, column=col, value=valor)
    celula.font = Font(bold=True)
    celula.fill = cinza
for col in range(1, 15):
    celula = abril.cell(row=3, column=col)
    celula.fill = cinza
abril.freeze_panes = "A4"
#aplicar estilo moeda
for row in range(3, 51):
    celula_coluna_2 = abril.cell(row=row, column=2)
    celula_coluna_7 = abril.cell(row=row, column=7)
    celula_coluna_12 = abril.cell(row=row, column=12)
    celula_coluna_2.style = moeda
    celula_coluna_7.style = moeda
    celula_coluna_12.style = moeda
#aplicar bordas
for i in range(1, num_linhas + 1):
    for j in range(1, num_colunas + 1):
        cell = abril.cell(row=i, column=j)
        cell.border = contorno
#formatar colunas
abril.column_dimensions[openpyxl.utils.get_column_letter(divisoria)].width = 0.5
abril.column_dimensions[openpyxl.utils.get_column_letter(divisoria2)].width = 0.5
abril.column_dimensions[openpyxl.utils.get_column_letter(compra1)].width = 30
abril.column_dimensions[openpyxl.utils.get_column_letter(compra2)].width = 30
abril.column_dimensions[openpyxl.utils.get_column_letter(compra3)].width = 30
abril.column_dimensions[openpyxl.utils.get_column_letter(loja1)].width = 30
abril.column_dimensions[openpyxl.utils.get_column_letter(loja2)].width = 30
abril.column_dimensions[openpyxl.utils.get_column_letter(loja3)].width = 30
abril.column_dimensions[openpyxl.utils.get_column_letter(preco1)].width = 11
abril.column_dimensions[openpyxl.utils.get_column_letter(preco2)].width = 11
abril.column_dimensions[openpyxl.utils.get_column_letter(preco3)].width = 11
abril.column_dimensions[openpyxl.utils.get_column_letter(data1)].width = 5
abril.column_dimensions[openpyxl.utils.get_column_letter(data2)].width = 5
abril.column_dimensions[openpyxl.utils.get_column_letter(data3)].width = 5
for row in range(1, 51):
    celula = abril.cell(row=row, column=divisoria)
    celula.fill = preto
for row in range(1, 51):
    celula = abril.cell(row=row, column=divisoria2)
    celula.fill = preto
celula_soma_2 = abril.cell(row=3, column=2, value=soma_coluna_2)
celula_soma_7 = abril.cell(row=3, column=7, value=soma_coluna_7)
celula_soma_12 = abril.cell(row=3, column=12, value=soma_coluna_12)
celula_soma_2.font = Font(bold=True)
celula_soma_2.fill = cinza
celula_soma_7.font = Font(bold=True)
celula_soma_7.fill = cinza
celula_soma_12.font = Font(bold=True)
celula_soma_12.fill = cinza
abril.cell(row=3, column=1, value="Total").font = Font(bold=True)
abril.cell(row=3, column=6, value="Total").font = Font(bold=True)
abril.cell(row=3, column=11, value="Total").font =Font(bold=True)
#página 5 (maio)
maio = planilha.create_sheet('Maio')
maio.merge_cells('A1:D1')
maio['A1'] = 'Cartão de crédito'
maio['A1'].alignment = Alignment(horizontal='center', vertical='center')
maio['A1'].font = Font(color="FFFFFF", bold=True, size=14)
maio['A1'].fill = preto
maio.merge_cells('F1:I1')
maio['F1'] = 'Vale alimentação'
maio['F1'].alignment = Alignment(horizontal='center', vertical='center')
maio['F1'].font = Font(color="FFFFFF", bold=True, size=14)
maio['f1'].fill = preto
maio.merge_cells('K1:N1')
maio['K1'] = 'Vale mobilidade'
maio['K1'].alignment = Alignment(horizontal='center', vertical='center')
maio['K1'].font = Font(color="FFFFFF", bold=True, size=14)
maio['K1'].fill = preto
cabecalho = ["Compra", "Preço", "Data", "Loja", " ", "Compra", "Preço", "Data", "Loja", " ", "Compra", "Preço", "Data", "Loja"]
for col, valor in enumerate(cabecalho, start=1):
    celula = maio.cell(row=2, column=col, value=valor)
    celula.font = Font(bold=True)
    celula.fill = cinza
for col in range(1, 15):
    celula = maio.cell(row=3, column=col)
    celula.fill = cinza
maio.freeze_panes = "A4"
#aplicar estilo moeda
for row in range(3, 51):
    celula_coluna_2 = maio.cell(row=row, column=2)
    celula_coluna_7 = maio.cell(row=row, column=7)
    celula_coluna_12 = maio.cell(row=row, column=12)
    celula_coluna_2.style = moeda
    celula_coluna_7.style = moeda
    celula_coluna_12.style = moeda
#aplicar bordas
for i in range(1, num_linhas + 1):
    for j in range(1, num_colunas + 1):
        cell = maio.cell(row=i, column=j)
        cell.border = contorno
#formatar colunas
maio.column_dimensions[openpyxl.utils.get_column_letter(divisoria)].width = 0.5
maio.column_dimensions[openpyxl.utils.get_column_letter(divisoria2)].width = 0.5
maio.column_dimensions[openpyxl.utils.get_column_letter(compra1)].width = 30
maio.column_dimensions[openpyxl.utils.get_column_letter(compra2)].width = 30
maio.column_dimensions[openpyxl.utils.get_column_letter(compra3)].width = 30
maio.column_dimensions[openpyxl.utils.get_column_letter(loja1)].width = 30
maio.column_dimensions[openpyxl.utils.get_column_letter(loja2)].width = 30
maio.column_dimensions[openpyxl.utils.get_column_letter(loja3)].width = 30
maio.column_dimensions[openpyxl.utils.get_column_letter(preco1)].width = 11
maio.column_dimensions[openpyxl.utils.get_column_letter(preco2)].width = 11
maio.column_dimensions[openpyxl.utils.get_column_letter(preco3)].width = 11
maio.column_dimensions[openpyxl.utils.get_column_letter(data1)].width = 5
maio.column_dimensions[openpyxl.utils.get_column_letter(data2)].width = 5
maio.column_dimensions[openpyxl.utils.get_column_letter(data3)].width = 5
for row in range(1, 51):
    celula = maio.cell(row=row, column=divisoria)
    celula.fill = preto
for row in range(1, 51):
    celula = maio.cell(row=row, column=divisoria2)
    celula.fill = preto
celula_soma_2 = maio.cell(row=3, column=2, value=soma_coluna_2)
celula_soma_7 = maio.cell(row=3, column=7, value=soma_coluna_7)
celula_soma_12 = maio.cell(row=3, column=12, value=soma_coluna_12)
celula_soma_2.font = Font(bold=True)
celula_soma_2.fill = cinza
celula_soma_7.font = Font(bold=True)
celula_soma_7.fill = cinza
celula_soma_12.font = Font(bold=True)
celula_soma_12.fill = cinza
maio.cell(row=3, column=1, value="Total").font = Font(bold=True)
maio.cell(row=3, column=6, value="Total").font = Font(bold=True)
maio.cell(row=3, column=11, value="Total").font =Font(bold=True)
#página 6 (junho)
junho = planilha.create_sheet('Junho')
junho.merge_cells('A1:D1')
junho['A1'] = 'Cartão de crédito'
junho['A1'].alignment = Alignment(horizontal='center', vertical='center')
junho['A1'].font = Font(color="FFFFFF", bold=True, size=14)
junho['A1'].fill = preto
junho.merge_cells('F1:I1')
junho['F1'] = 'Vale alimentação'
junho['F1'].alignment = Alignment(horizontal='center', vertical='center')
junho['F1'].font = Font(color="FFFFFF", bold=True, size=14)
junho['f1'].fill = preto
junho.merge_cells('K1:N1')
junho['K1'] = 'Vale mobilidade'
junho['K1'].alignment = Alignment(horizontal='center', vertical='center')
junho['K1'].font = Font(color="FFFFFF", bold=True, size=14)
junho['K1'].fill = preto
cabecalho = ["Compra", "Preço", "Data", "Loja", " ", "Compra", "Preço", "Data", "Loja", " ", "Compra", "Preço", "Data", "Loja"]
for col, valor in enumerate(cabecalho, start=1):
    celula = junho.cell(row=2, column=col, value=valor)
    celula.font = Font(bold=True)
    celula.fill = cinza
for col in range(1, 15):
    celula = junho.cell(row=3, column=col)
    celula.fill = cinza
junho.freeze_panes = "A4"
#aplicar estilo moeda
for row in range(3, 51):
    celula_coluna_2 = junho.cell(row=row, column=2)
    celula_coluna_7 = junho.cell(row=row, column=7)
    celula_coluna_12 = junho.cell(row=row, column=12)
    celula_coluna_2.style = moeda
    celula_coluna_7.style = moeda
    celula_coluna_12.style = moeda
#aplicar bordas
for i in range(1, num_linhas + 1):
    for j in range(1, num_colunas + 1):
        cell = junho.cell(row=i, column=j)
        cell.border = contorno
#formatar colunas
junho.column_dimensions[openpyxl.utils.get_column_letter(divisoria)].width = 0.5
junho.column_dimensions[openpyxl.utils.get_column_letter(divisoria2)].width = 0.5
junho.column_dimensions[openpyxl.utils.get_column_letter(compra1)].width = 30
junho.column_dimensions[openpyxl.utils.get_column_letter(compra2)].width = 30
junho.column_dimensions[openpyxl.utils.get_column_letter(compra3)].width = 30
junho.column_dimensions[openpyxl.utils.get_column_letter(loja1)].width = 30
junho.column_dimensions[openpyxl.utils.get_column_letter(loja2)].width = 30
junho.column_dimensions[openpyxl.utils.get_column_letter(loja3)].width = 30
junho.column_dimensions[openpyxl.utils.get_column_letter(preco1)].width = 11
junho.column_dimensions[openpyxl.utils.get_column_letter(preco2)].width = 11
junho.column_dimensions[openpyxl.utils.get_column_letter(preco3)].width = 11
junho.column_dimensions[openpyxl.utils.get_column_letter(data1)].width = 5
junho.column_dimensions[openpyxl.utils.get_column_letter(data2)].width = 5
junho.column_dimensions[openpyxl.utils.get_column_letter(data3)].width = 5
for row in range(1, 51):
    celula = junho.cell(row=row, column=divisoria)
    celula.fill = preto
for row in range(1, 51):
    celula = junho.cell(row=row, column=divisoria2)
    celula.fill = preto
celula_soma_2 = junho.cell(row=3, column=2, value=soma_coluna_2)
celula_soma_7 = junho.cell(row=3, column=7, value=soma_coluna_7)
celula_soma_12 = junho.cell(row=3, column=12, value=soma_coluna_12)
celula_soma_2.font = Font(bold=True)
celula_soma_2.fill = cinza
celula_soma_7.font = Font(bold=True)
celula_soma_7.fill = cinza
celula_soma_12.font = Font(bold=True)
celula_soma_12.fill = cinza
junho.cell(row=3, column=1, value="Total").font = Font(bold=True)
junho.cell(row=3, column=6, value="Total").font = Font(bold=True)
junho.cell(row=3, column=11, value="Total").font =Font(bold=True)
#página 7 (julho)
julho = planilha.create_sheet('Julho')
julho.merge_cells('A1:D1')
julho['A1'] = 'Cartão de crédito'
julho['A1'].alignment = Alignment(horizontal='center', vertical='center')
julho['A1'].font = Font(color="FFFFFF", bold=True, size=14)
julho['A1'].fill = preto
julho.merge_cells('F1:I1')
julho['F1'] = 'Vale alimentação'
julho['F1'].alignment = Alignment(horizontal='center', vertical='center')
julho['F1'].font = Font(color="FFFFFF", bold=True, size=14)
julho['f1'].fill = preto
julho.merge_cells('K1:N1')
julho['K1'] = 'Vale mobilidade'
julho['K1'].alignment = Alignment(horizontal='center', vertical='center')
julho['K1'].font = Font(color="FFFFFF", bold=True, size=14)
julho['K1'].fill = preto
cabecalho = ["Compra", "Preço", "Data", "Loja", " ", "Compra", "Preço", "Data", "Loja", " ", "Compra", "Preço", "Data", "Loja"]
for col, valor in enumerate(cabecalho, start=1):
    celula = julho.cell(row=2, column=col, value=valor)
    celula.font = Font(bold=True)
    celula.fill = cinza
for col in range(1, 15):
    celula = julho.cell(row=3, column=col)
    celula.fill = cinza
julho.freeze_panes = "A4"
#aplicar estilo moeda
for row in range(3, 51):
    celula_coluna_2 = julho.cell(row=row, column=2)
    celula_coluna_7 = julho.cell(row=row, column=7)
    celula_coluna_12 = julho.cell(row=row, column=12)
    celula_coluna_2.style = moeda
    celula_coluna_7.style = moeda
    celula_coluna_12.style = moeda
#aplicar bordas
for i in range(1, num_linhas + 1):
    for j in range(1, num_colunas + 1):
        cell = julho.cell(row=i, column=j)
        cell.border = contorno
#formatar colunas
julho.column_dimensions[openpyxl.utils.get_column_letter(divisoria)].width = 0.5
julho.column_dimensions[openpyxl.utils.get_column_letter(divisoria2)].width = 0.5
julho.column_dimensions[openpyxl.utils.get_column_letter(compra1)].width = 30
julho.column_dimensions[openpyxl.utils.get_column_letter(compra2)].width = 30
julho.column_dimensions[openpyxl.utils.get_column_letter(compra3)].width = 30
julho.column_dimensions[openpyxl.utils.get_column_letter(loja1)].width = 30
julho.column_dimensions[openpyxl.utils.get_column_letter(loja2)].width = 30
julho.column_dimensions[openpyxl.utils.get_column_letter(loja3)].width = 30
julho.column_dimensions[openpyxl.utils.get_column_letter(preco1)].width = 11
julho.column_dimensions[openpyxl.utils.get_column_letter(preco2)].width = 11
julho.column_dimensions[openpyxl.utils.get_column_letter(preco3)].width = 11
julho.column_dimensions[openpyxl.utils.get_column_letter(data1)].width = 5
julho.column_dimensions[openpyxl.utils.get_column_letter(data2)].width = 5
julho.column_dimensions[openpyxl.utils.get_column_letter(data3)].width = 5
for row in range(1, 51):
    celula = julho.cell(row=row, column=divisoria)
    celula.fill = preto
for row in range(1, 51):
    celula = julho.cell(row=row, column=divisoria2)
    celula.fill = preto
celula_soma_2 = julho.cell(row=3, column=2, value=soma_coluna_2)
celula_soma_7 = julho.cell(row=3, column=7, value=soma_coluna_7)
celula_soma_12 = julho.cell(row=3, column=12, value=soma_coluna_12)
celula_soma_2.font = Font(bold=True)
celula_soma_2.fill = cinza
celula_soma_7.font = Font(bold=True)
celula_soma_7.fill = cinza
celula_soma_12.font = Font(bold=True)
celula_soma_12.fill = cinza
julho.cell(row=3, column=1, value="Total").font = Font(bold=True)
julho.cell(row=3, column=6, value="Total").font = Font(bold=True)
julho.cell(row=3, column=11, value="Total").font =Font(bold=True)
#página 8 (agosto)
agosto = planilha.create_sheet('Agosto')
agosto.merge_cells('A1:D1')
agosto['A1'] = 'Cartão de crédito'
agosto['A1'].alignment = Alignment(horizontal='center', vertical='center')
agosto['A1'].font = Font(color="FFFFFF", bold=True, size=14)
agosto['A1'].fill = preto
agosto.merge_cells('F1:I1')
agosto['F1'] = 'Vale alimentação'
agosto['F1'].alignment = Alignment(horizontal='center', vertical='center')
agosto['F1'].font = Font(color="FFFFFF", bold=True, size=14)
agosto['f1'].fill = preto
agosto.merge_cells('K1:N1')
agosto['K1'] = 'Vale mobilidade'
agosto['K1'].alignment = Alignment(horizontal='center', vertical='center')
agosto['K1'].font = Font(color="FFFFFF", bold=True, size=14)
agosto['K1'].fill = preto
cabecalho = ["Compra", "Preço", "Data", "Loja", " ", "Compra", "Preço", "Data", "Loja", " ", "Compra", "Preço", "Data", "Loja"]
for col, valor in enumerate(cabecalho, start=1):
    celula = agosto.cell(row=2, column=col, value=valor)
    celula.font = Font(bold=True)
    celula.fill = cinza
for col in range(1, 15):
    celula = agosto.cell(row=3, column=col)
    celula.fill = cinza
agosto.freeze_panes = "A4"
#aplicar estilo moeda
for row in range(3, 51):
    celula_coluna_2 = agosto.cell(row=row, column=2)
    celula_coluna_7 = agosto.cell(row=row, column=7)
    celula_coluna_12 = agosto.cell(row=row, column=12)
    celula_coluna_2.style = moeda
    celula_coluna_7.style = moeda
    celula_coluna_12.style = moeda
#aplicar bordas
for i in range(1, num_linhas + 1):
    for j in range(1, num_colunas + 1):
        cell = agosto.cell(row=i, column=j)
        cell.border = contorno
#formatar colunas
agosto.column_dimensions[openpyxl.utils.get_column_letter(divisoria)].width = 0.5
agosto.column_dimensions[openpyxl.utils.get_column_letter(divisoria2)].width = 0.5
agosto.column_dimensions[openpyxl.utils.get_column_letter(compra1)].width = 30
agosto.column_dimensions[openpyxl.utils.get_column_letter(compra2)].width = 30
agosto.column_dimensions[openpyxl.utils.get_column_letter(compra3)].width = 30
agosto.column_dimensions[openpyxl.utils.get_column_letter(loja1)].width = 30
agosto.column_dimensions[openpyxl.utils.get_column_letter(loja2)].width = 30
agosto.column_dimensions[openpyxl.utils.get_column_letter(loja3)].width = 30
agosto.column_dimensions[openpyxl.utils.get_column_letter(preco1)].width = 11
agosto.column_dimensions[openpyxl.utils.get_column_letter(preco2)].width = 11
agosto.column_dimensions[openpyxl.utils.get_column_letter(preco3)].width = 11
agosto.column_dimensions[openpyxl.utils.get_column_letter(data1)].width = 5
agosto.column_dimensions[openpyxl.utils.get_column_letter(data2)].width = 5
agosto.column_dimensions[openpyxl.utils.get_column_letter(data3)].width = 5
for row in range(1, 51):
    celula = agosto.cell(row=row, column=divisoria)
    celula.fill = preto
for row in range(1, 51):
    celula = agosto.cell(row=row, column=divisoria2)
    celula.fill = preto
celula_soma_2 = agosto.cell(row=3, column=2, value=soma_coluna_2)
celula_soma_7 = agosto.cell(row=3, column=7, value=soma_coluna_7)
celula_soma_12 = agosto.cell(row=3, column=12, value=soma_coluna_12)
celula_soma_2.font = Font(bold=True)
celula_soma_2.fill = cinza
celula_soma_7.font = Font(bold=True)
celula_soma_7.fill = cinza
celula_soma_12.font = Font(bold=True)
celula_soma_12.fill = cinza
agosto.cell(row=3, column=1, value="Total").font = Font(bold=True)
agosto.cell(row=3, column=6, value="Total").font = Font(bold=True)
agosto.cell(row=3, column=11, value="Total").font =Font(bold=True)
#página 9 (setembro)
setembro = planilha.create_sheet('Setembro')
setembro.merge_cells('A1:D1')
setembro['A1'] = 'Cartão de crédito'
setembro['A1'].alignment = Alignment(horizontal='center', vertical='center')
setembro['A1'].font = Font(color="FFFFFF", bold=True, size=14)
setembro['A1'].fill = preto
setembro.merge_cells('F1:I1')
setembro['F1'] = 'Vale alimentação'
setembro['F1'].alignment = Alignment(horizontal='center', vertical='center')
setembro['F1'].font = Font(color="FFFFFF", bold=True, size=14)
setembro['f1'].fill = preto
setembro.merge_cells('K1:N1')
setembro['K1'] = 'Vale mobilidade'
setembro['K1'].alignment = Alignment(horizontal='center', vertical='center')
setembro['K1'].font = Font(color="FFFFFF", bold=True, size=14)
setembro['K1'].fill = preto
cabecalho = ["Compra", "Preço", "Data", "Loja", " ", "Compra", "Preço", "Data", "Loja", " ", "Compra", "Preço", "Data", "Loja"]
for col, valor in enumerate(cabecalho, start=1):
    celula = setembro.cell(row=2, column=col, value=valor)
    celula.font = Font(bold=True)
    celula.fill = cinza
for col in range(1, 15):
    celula = setembro.cell(row=3, column=col)
    celula.fill = cinza
setembro.freeze_panes = "A4"
#aplicar estilo moeda
for row in range(3, 51):
    celula_coluna_2 = setembro.cell(row=row, column=2)
    celula_coluna_7 = setembro.cell(row=row, column=7)
    celula_coluna_12 = setembro.cell(row=row, column=12)
    celula_coluna_2.style = moeda
    celula_coluna_7.style = moeda
    celula_coluna_12.style = moeda
#aplicar bordas
for i in range(1, num_linhas + 1):
    for j in range(1, num_colunas + 1):
        cell = setembro.cell(row=i, column=j)
        cell.border = contorno
#formatar colunas
setembro.column_dimensions[openpyxl.utils.get_column_letter(divisoria)].width = 0.5
setembro.column_dimensions[openpyxl.utils.get_column_letter(divisoria2)].width = 0.5
setembro.column_dimensions[openpyxl.utils.get_column_letter(compra1)].width = 30
setembro.column_dimensions[openpyxl.utils.get_column_letter(compra2)].width = 30
setembro.column_dimensions[openpyxl.utils.get_column_letter(compra3)].width = 30
setembro.column_dimensions[openpyxl.utils.get_column_letter(loja1)].width = 30
setembro.column_dimensions[openpyxl.utils.get_column_letter(loja2)].width = 30
setembro.column_dimensions[openpyxl.utils.get_column_letter(loja3)].width = 30
setembro.column_dimensions[openpyxl.utils.get_column_letter(preco1)].width = 11
setembro.column_dimensions[openpyxl.utils.get_column_letter(preco2)].width = 11
setembro.column_dimensions[openpyxl.utils.get_column_letter(preco3)].width = 11
setembro.column_dimensions[openpyxl.utils.get_column_letter(data1)].width = 5
setembro.column_dimensions[openpyxl.utils.get_column_letter(data2)].width = 5
setembro.column_dimensions[openpyxl.utils.get_column_letter(data3)].width = 5
for row in range(1, 51):
    celula = setembro.cell(row=row, column=divisoria)
    celula.fill = preto
for row in range(1, 51):
    celula = setembro.cell(row=row, column=divisoria2)
    celula.fill = preto
celula_soma_2 = setembro.cell(row=3, column=2, value=soma_coluna_2)
celula_soma_7 = setembro.cell(row=3, column=7, value=soma_coluna_7)
celula_soma_12 = setembro.cell(row=3, column=12, value=soma_coluna_12)
celula_soma_2.font = Font(bold=True)
celula_soma_2.fill = cinza
celula_soma_7.font = Font(bold=True)
celula_soma_7.fill = cinza
celula_soma_12.font = Font(bold=True)
celula_soma_12.fill = cinza
setembro.cell(row=3, column=1, value="Total").font = Font(bold=True)
setembro.cell(row=3, column=6, value="Total").font = Font(bold=True)
setembro.cell(row=3, column=11, value="Total").font =Font(bold=True)
#página 10 (outubro)
outubro = planilha.create_sheet('Outubro')
outubro.merge_cells('A1:D1')
outubro['A1'] = 'Cartão de crédito'
outubro['A1'].alignment = Alignment(horizontal='center', vertical='center')
outubro['A1'].font = Font(color="FFFFFF", bold=True, size=14)
outubro['A1'].fill = preto
outubro.merge_cells('F1:I1')
outubro['F1'] = 'Vale alimentação'
outubro['F1'].alignment = Alignment(horizontal='center', vertical='center')
outubro['F1'].font = Font(color="FFFFFF", bold=True, size=14)
outubro['f1'].fill = preto
outubro.merge_cells('K1:N1')
outubro['K1'] = 'Vale mobilidade'
outubro['K1'].alignment = Alignment(horizontal='center', vertical='center')
outubro['K1'].font = Font(color="FFFFFF", bold=True, size=14)
outubro['K1'].fill = preto
cabecalho = ["Compra", "Preço", "Data", "Loja", " ", "Compra", "Preço", "Data", "Loja", " ", "Compra", "Preço", "Data", "Loja"]
for col, valor in enumerate(cabecalho, start=1):
    celula = outubro.cell(row=2, column=col, value=valor)
    celula.font = Font(bold=True)
    celula.fill = cinza
for col in range(1, 15):
    celula = outubro.cell(row=3, column=col)
    celula.fill = cinza
outubro.freeze_panes = "A4"
#aplicar estilo moeda
for row in range(3, 51):
    celula_coluna_2 = outubro.cell(row=row, column=2)
    celula_coluna_7 = outubro.cell(row=row, column=7)
    celula_coluna_12 = outubro.cell(row=row, column=12)
    celula_coluna_2.style = moeda
    celula_coluna_7.style = moeda
    celula_coluna_12.style = moeda
#aplicar bordas
for i in range(1, num_linhas + 1):
    for j in range(1, num_colunas + 1):
        cell = outubro.cell(row=i, column=j)
        cell.border = contorno
#formatar colunas
outubro.column_dimensions[openpyxl.utils.get_column_letter(divisoria)].width = 0.5
outubro.column_dimensions[openpyxl.utils.get_column_letter(divisoria2)].width = 0.5
outubro.column_dimensions[openpyxl.utils.get_column_letter(compra1)].width = 30
outubro.column_dimensions[openpyxl.utils.get_column_letter(compra2)].width = 30
outubro.column_dimensions[openpyxl.utils.get_column_letter(compra3)].width = 30
outubro.column_dimensions[openpyxl.utils.get_column_letter(loja1)].width = 30
outubro.column_dimensions[openpyxl.utils.get_column_letter(loja2)].width = 30
outubro.column_dimensions[openpyxl.utils.get_column_letter(loja3)].width = 30
outubro.column_dimensions[openpyxl.utils.get_column_letter(preco1)].width = 11
outubro.column_dimensions[openpyxl.utils.get_column_letter(preco2)].width = 11
outubro.column_dimensions[openpyxl.utils.get_column_letter(preco3)].width = 11
outubro.column_dimensions[openpyxl.utils.get_column_letter(data1)].width = 5
outubro.column_dimensions[openpyxl.utils.get_column_letter(data2)].width = 5
outubro.column_dimensions[openpyxl.utils.get_column_letter(data3)].width = 5
for row in range(1, 51):
    celula = outubro.cell(row=row, column=divisoria)
    celula.fill = preto
for row in range(1, 51):
    celula = outubro.cell(row=row, column=divisoria2)
    celula.fill = preto
celula_soma_2 = outubro.cell(row=3, column=2, value=soma_coluna_2)
celula_soma_7 = outubro.cell(row=3, column=7, value=soma_coluna_7)
celula_soma_12 = outubro.cell(row=3, column=12, value=soma_coluna_12)
celula_soma_2.font = Font(bold=True)
celula_soma_2.fill = cinza
celula_soma_7.font = Font(bold=True)
celula_soma_7.fill = cinza
celula_soma_12.font = Font(bold=True)
celula_soma_12.fill = cinza
outubro.cell(row=3, column=1, value="Total").font = Font(bold=True)
outubro.cell(row=3, column=6, value="Total").font = Font(bold=True)
outubro.cell(row=3, column=11, value="Total").font =Font(bold=True)
#página 11 (novembro)
novembro = planilha.create_sheet('Novembro')
novembro.merge_cells('A1:D1')
novembro['A1'] = 'Cartão de crédito'
novembro['A1'].alignment = Alignment(horizontal='center', vertical='center')
novembro['A1'].font = Font(color="FFFFFF", bold=True, size=14)
novembro['A1'].fill = preto
novembro.merge_cells('F1:I1')
novembro['F1'] = 'Vale alimentação'
novembro['F1'].alignment = Alignment(horizontal='center', vertical='center')
novembro['F1'].font = Font(color="FFFFFF", bold=True, size=14)
novembro['f1'].fill = preto
novembro.merge_cells('K1:N1')
novembro['K1'] = 'Vale mobilidade'
novembro['K1'].alignment = Alignment(horizontal='center', vertical='center')
novembro['K1'].font = Font(color="FFFFFF", bold=True, size=14)
novembro['K1'].fill = preto
cabecalho = ["Compra", "Preço", "Data", "Loja", " ", "Compra", "Preço", "Data", "Loja", " ", "Compra", "Preço", "Data", "Loja"]
for col, valor in enumerate(cabecalho, start=1):
    celula = novembro.cell(row=2, column=col, value=valor)
    celula.font = Font(bold=True)
    celula.fill = cinza
for col in range(1, 15):
    celula = novembro.cell(row=3, column=col)
    celula.fill = cinza
novembro.freeze_panes = "A4"
#aplicar estilo moeda
for row in range(3, 51):
    celula_coluna_2 = novembro.cell(row=row, column=2)
    celula_coluna_7 = novembro.cell(row=row, column=7)
    celula_coluna_12 = novembro.cell(row=row, column=12)
    celula_coluna_2.style = moeda
    celula_coluna_7.style = moeda
    celula_coluna_12.style = moeda
#aplicar bordas
for i in range(1, num_linhas + 1):
    for j in range(1, num_colunas + 1):
        cell = novembro.cell(row=i, column=j)
        cell.border = contorno
#formatar colunas
novembro.column_dimensions[openpyxl.utils.get_column_letter(divisoria)].width = 0.5
novembro.column_dimensions[openpyxl.utils.get_column_letter(divisoria2)].width = 0.5
novembro.column_dimensions[openpyxl.utils.get_column_letter(compra1)].width = 30
novembro.column_dimensions[openpyxl.utils.get_column_letter(compra2)].width = 30
novembro.column_dimensions[openpyxl.utils.get_column_letter(compra3)].width = 30
novembro.column_dimensions[openpyxl.utils.get_column_letter(loja1)].width = 30
novembro.column_dimensions[openpyxl.utils.get_column_letter(loja2)].width = 30
novembro.column_dimensions[openpyxl.utils.get_column_letter(loja3)].width = 30
novembro.column_dimensions[openpyxl.utils.get_column_letter(preco1)].width = 11
novembro.column_dimensions[openpyxl.utils.get_column_letter(preco2)].width = 11
novembro.column_dimensions[openpyxl.utils.get_column_letter(preco3)].width = 11
novembro.column_dimensions[openpyxl.utils.get_column_letter(data1)].width = 5
novembro.column_dimensions[openpyxl.utils.get_column_letter(data2)].width = 5
novembro.column_dimensions[openpyxl.utils.get_column_letter(data3)].width = 5
for row in range(1, 51):
    celula = novembro.cell(row=row, column=divisoria)
    celula.fill = preto
for row in range(1, 51):
    celula = novembro.cell(row=row, column=divisoria2)
    celula.fill = preto
celula_soma_2 = novembro.cell(row=3, column=2, value=soma_coluna_2)
celula_soma_7 = novembro.cell(row=3, column=7, value=soma_coluna_7)
celula_soma_12 = novembro.cell(row=3, column=12, value=soma_coluna_12)
celula_soma_2.font = Font(bold=True)
celula_soma_2.fill = cinza
celula_soma_7.font = Font(bold=True)
celula_soma_7.fill = cinza
celula_soma_12.font = Font(bold=True)
celula_soma_12.fill = cinza
novembro.cell(row=3, column=1, value="Total").font = Font(bold=True)
novembro.cell(row=3, column=6, value="Total").font = Font(bold=True)
novembro.cell(row=3, column=11, value="Total").font =Font(bold=True)
novembro.cell(row=50, column=11, value="Total").font =Font(bold=True)
#página 12 (dezembro)
dezembro = planilha.create_sheet('Dezembro')
dezembro.merge_cells('A1:D1')
dezembro['A1'] = 'Cartão de crédito'
dezembro['A1'].alignment = Alignment(horizontal='center', vertical='center')
dezembro['A1'].font = Font(color="FFFFFF", bold=True, size=14)
dezembro['A1'].fill = preto
dezembro.merge_cells('F1:I1')
dezembro['F1'] = 'Vale alimentação'
dezembro['F1'].alignment = Alignment(horizontal='center', vertical='center')
dezembro['F1'].font = Font(color="FFFFFF", bold=True, size=14)
dezembro['f1'].fill = preto
dezembro.merge_cells('K1:N1')
dezembro['K1'] = 'Vale mobilidade'
dezembro['K1'].alignment = Alignment(horizontal='center', vertical='center')
dezembro['K1'].font = Font(color="FFFFFF", bold=True, size=14)
dezembro['K1'].fill = preto
cabecalho = ["Compra", "Preço", "Data", "Loja", " ", "Compra", "Preço", "Data", "Loja", " ", "Compra", "Preço", "Data", "Loja"]
for col, valor in enumerate(cabecalho, start=1):
    celula = dezembro.cell(row=2, column=col, value=valor)
    celula.font = Font(bold=True)
    celula.fill = cinza
for col in range(1, 15):
    celula = dezembro.cell(row=3, column=col)
    celula.fill = cinza
dezembro.freeze_panes = "A4"
#aplicar estilo moeda
for row in range(3, 51):
    celula_coluna_2 = dezembro.cell(row=row, column=2)
    celula_coluna_7 = dezembro.cell(row=row, column=7)
    celula_coluna_12 = dezembro.cell(row=row, column=12)
    celula_coluna_2.style = moeda
    celula_coluna_7.style = moeda
    celula_coluna_12.style = moeda
#aplicar bordas
for i in range(1, num_linhas + 1):
    for j in range(1, num_colunas + 1):
        cell = dezembro.cell(row=i, column=j)
        cell.border = contorno
#formatar colunas
dezembro.column_dimensions[openpyxl.utils.get_column_letter(divisoria)].width = 0.5
dezembro.column_dimensions[openpyxl.utils.get_column_letter(divisoria2)].width = 0.5
dezembro.column_dimensions[openpyxl.utils.get_column_letter(compra1)].width = 30
dezembro.column_dimensions[openpyxl.utils.get_column_letter(compra2)].width = 30
dezembro.column_dimensions[openpyxl.utils.get_column_letter(compra3)].width = 30
dezembro.column_dimensions[openpyxl.utils.get_column_letter(loja1)].width = 30
dezembro.column_dimensions[openpyxl.utils.get_column_letter(loja2)].width = 30
dezembro.column_dimensions[openpyxl.utils.get_column_letter(loja3)].width = 30
dezembro.column_dimensions[openpyxl.utils.get_column_letter(preco1)].width = 11
dezembro.column_dimensions[openpyxl.utils.get_column_letter(preco2)].width = 11
dezembro.column_dimensions[openpyxl.utils.get_column_letter(preco3)].width = 11
dezembro.column_dimensions[openpyxl.utils.get_column_letter(data1)].width = 5
dezembro.column_dimensions[openpyxl.utils.get_column_letter(data2)].width = 5
dezembro.column_dimensions[openpyxl.utils.get_column_letter(data3)].width = 5
for row in range(1, 51):
    celula = dezembro.cell(row=row, column=divisoria)
    celula.fill = preto
for row in range(1, 51):
    celula = dezembro.cell(row=row, column=divisoria2)
    celula.fill = preto
celula_soma_2 = dezembro.cell(row=3, column=2, value=soma_coluna_2)
celula_soma_7 = dezembro.cell(row=3, column=7, value=soma_coluna_7)
celula_soma_12 = dezembro.cell(row=3, column=12, value=soma_coluna_12)
celula_soma_2.font = Font(bold=True)
celula_soma_2.fill = cinza
celula_soma_7.font = Font(bold=True)
celula_soma_7.fill = cinza
celula_soma_12.font = Font(bold=True)
celula_soma_12.fill = cinza
dezembro.cell(row=3, column=1, value="Total").font = Font(bold=True)
dezembro.cell(row=3, column=6, value="Total").font = Font(bold=True)
dezembro.cell(row=3, column=11, value="Total").font =Font(bold=True)

#salvar planilha
planilha.save(r"C:\planilhav3.xlsx")
print("Planilha de gastos criada com sucesso.")
